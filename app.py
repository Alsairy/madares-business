from flask import Flask, request, jsonify, render_template_string, send_file
from flask_cors import CORS
import sqlite3
import json
import csv
import io
import os
import uuid
from datetime import datetime
import base64
from werkzeug.utils import secure_filename
import pytesseract
from PIL import Image
import PyPDF2
import docx
import openpyxl

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx'}
DATABASE_PATH = 'madares_complete.db'

# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
for doc_type in ['property_deed', 'ownership_documents', 'construction_plans', 'financial_documents', 'legal_documents', 'inspection_reports']:
    os.makedirs(os.path.join(UPLOAD_FOLDER, doc_type), exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_file(file_path):
    """Extract text from various file formats using OCR and document parsing"""
    try:
        file_ext = file_path.rsplit('.', 1)[1].lower()
        
        if file_ext in ['png', 'jpg', 'jpeg', 'gif']:
            # OCR for images
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text.strip()
            
        elif file_ext == 'pdf':
            # Extract text from PDF
            text = ""
            try:
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text += page.extract_text()
                if not text.strip():
                    # If no text extracted, try OCR on PDF pages
                    from pdf2image import convert_from_path
                    pages = convert_from_path(file_path)
                    for page in pages:
                        text += pytesseract.image_to_string(page)
            except:
                # Fallback OCR
                from pdf2image import convert_from_path
                pages = convert_from_path(file_path)
                for page in pages:
                    text += pytesseract.image_to_string(page)
            return text.strip()
            
        elif file_ext in ['doc', 'docx']:
            # Extract text from Word documents
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\\n"
            return text.strip()
            
        elif file_ext in ['xls', 'xlsx']:
            # Extract text from Excel files
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text += str(cell.value) + " "
                    text += "\\n"
            return text.strip()
            
        else:
            # Plain text files
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read().strip()
                
    except Exception as e:
        return f"Error extracting text: {str(e)}"

# Initialize database
def init_db():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    # Create comprehensive assets table with all 58 MOE fields
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id TEXT UNIQUE,
            
            -- Section 1: Asset Identification & Status (6 fields)
            asset_name TEXT,
            asset_type TEXT,
            asset_status TEXT,
            asset_category TEXT,
            priority_level TEXT,
            strategic_importance TEXT,
            
            -- Section 2: Planning & Need Assessment (4 fields)
            planning_status TEXT,
            need_assessment TEXT,
            feasibility_study TEXT,
            approval_status TEXT,
            
            -- Section 3: Location Attractiveness (3 fields)
            location_score INTEGER,
            accessibility_rating TEXT,
            infrastructure_quality TEXT,
            
            -- Section 4: Investment Proposal & Obstacles (3 fields)
            investment_proposal TEXT,
            identified_obstacles TEXT,
            mitigation_strategies TEXT,
            
            -- Section 5: Financial Obligations & Covenants (3 fields)
            financial_obligations TEXT,
            loan_covenants TEXT,
            payment_schedule TEXT,
            
            -- Section 6: Utilities Information (4 fields)
            electricity_connection TEXT,
            water_connection TEXT,
            sewage_connection TEXT,
            telecommunications TEXT,
            
            -- Section 7: Ownership Information (4 fields)
            ownership_type TEXT,
            owner_name TEXT,
            ownership_percentage REAL,
            ownership_documents TEXT,
            
            -- Section 8: Land & Plan Details (3 fields)
            land_area REAL,
            zoning_classification TEXT,
            land_use_permit TEXT,
            
            -- Section 9: Asset Area Details (5 fields)
            total_built_area REAL,
            usable_area REAL,
            common_area REAL,
            parking_area REAL,
            green_space_area REAL,
            
            -- Section 10: Construction Status (4 fields)
            construction_status TEXT,
            completion_percentage INTEGER,
            construction_start_date TEXT,
            expected_completion_date TEXT,
            
            -- Section 11: Physical Dimensions (4 fields)
            length_meters REAL,
            width_meters REAL,
            height_meters REAL,
            floor_count INTEGER,
            
            -- Section 12: Boundaries (8 fields)
            north_boundary TEXT,
            south_boundary TEXT,
            east_boundary TEXT,
            west_boundary TEXT,
            boundary_length_north REAL,
            boundary_length_south REAL,
            boundary_length_east REAL,
            boundary_length_west REAL,
            
            -- Section 13: Geographic Location (7 fields)
            region TEXT,
            city TEXT,
            district TEXT,
            street_address TEXT,
            latitude REAL,
            longitude REAL,
            elevation_meters REAL,
            
            -- Financial Information
            investment_value REAL,
            maintenance_cost REAL,
            
            -- Timestamps
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create files table for document management
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id INTEGER,
            file_name TEXT,
            file_path TEXT,
            file_type TEXT,
            document_category TEXT,
            file_size INTEGER,
            ocr_text TEXT,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (asset_id) REFERENCES assets (id)
        )
    ''')
    
    # Create workflows table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS workflows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workflow_id TEXT UNIQUE,
            title TEXT,
            description TEXT,
            priority TEXT,
            status TEXT,
            assigned_to TEXT,
            due_date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT UNIQUE,
            full_name TEXT,
            email TEXT,
            department TEXT,
            role TEXT,
            region TEXT,
            phone TEXT,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Insert sample data if tables are empty
    cursor.execute('SELECT COUNT(*) FROM assets')
    if cursor.fetchone()[0] == 0:
        sample_asset = (
            'AST-001', 'Riyadh Educational Complex', 'Educational', 'Active', 'Primary', 'High', 'Strategic',
            'Approved', 'Completed', 'Feasible', 'Approved',
            85, 'Excellent', 'High Quality',
            'Approved for 15M SAR investment', 'Traffic congestion during peak hours', 'Alternative access routes planned',
            'Standard loan terms', 'Quarterly payments required', 'Monthly installments',
            'Connected', 'Connected', 'Connected', 'Fiber optic available',
            'Government', 'Ministry of Education', 100.0, 'Complete ownership documents',
            5000.0, 'Educational', 'Approved',
            8500.0, 8000.0, 300.0, 200.0, 500.0,
            'In Progress', 75, '2023-01-15', '2024-06-30',
            120.0, 80.0, 25.0, 4,
            'Public road', 'Residential area', 'Commercial district', 'Park',
            120.0, 120.0, 80.0, 80.0,
            'Riyadh', 'Riyadh', 'Al-Malaz', '123 King Fahd Road',
            24.7136, 46.6753, 612.0,
            15000000, 50000
        )
        
        cursor.execute('''
            INSERT INTO assets (
                asset_id, asset_name, asset_type, asset_status, asset_category, priority_level, strategic_importance,
                planning_status, need_assessment, feasibility_study, approval_status,
                location_score, accessibility_rating, infrastructure_quality,
                investment_proposal, identified_obstacles, mitigation_strategies,
                financial_obligations, loan_covenants, payment_schedule,
                electricity_connection, water_connection, sewage_connection, telecommunications,
                ownership_type, owner_name, ownership_percentage, ownership_documents,
                land_area, zoning_classification, land_use_permit,
                total_built_area, usable_area, common_area, parking_area, green_space_area,
                construction_status, completion_percentage, construction_start_date, expected_completion_date,
                length_meters, width_meters, height_meters, floor_count,
                north_boundary, south_boundary, east_boundary, west_boundary,
                boundary_length_north, boundary_length_south, boundary_length_east, boundary_length_west,
                region, city, district, street_address, latitude, longitude, elevation_meters,
                investment_value, maintenance_cost
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', sample_asset)
    
    # Insert sample workflows
    cursor.execute('SELECT COUNT(*) FROM workflows')
    if cursor.fetchone()[0] == 0:
        sample_workflows = [
            ('WF-001', 'Asset Inspection - AST-001', 'Conduct quarterly inspection of Riyadh Educational Complex', 'High', 'In Progress', 'Ahmed Al-Rashid', '2024-02-15'),
            ('WF-002', 'Maintenance Planning - AST-001', 'Plan annual maintenance for educational complex', 'Medium', 'Not Started', 'Sara Al-Mahmoud', '2024-02-20')
        ]
        
        cursor.executemany('''
            INSERT INTO workflows (workflow_id, title, description, priority, status, assigned_to, due_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', sample_workflows)
    
    # Insert sample users
    cursor.execute('SELECT COUNT(*) FROM users')
    if cursor.fetchone()[0] == 0:
        sample_users = [
            ('USR-001', 'Ahmed Al-Rashid', 'ahmed.rashid@madares.sa', 'Asset Management', 'Senior Manager', 'Riyadh', '+966501234567'),
            ('USR-002', 'Sara Al-Mahmoud', 'sara.mahmoud@madares.sa', 'Operations', 'Project Coordinator', 'Jeddah', '+966509876543'),
            ('USR-003', 'Mohammed Al-Otaibi', 'mohammed.otaibi@madares.sa', 'Finance', 'Financial Analyst', 'Dammam', '+966505555555')
        ]
        
        cursor.executemany('''
            INSERT INTO users (user_id, full_name, email, department, role, region, phone)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', sample_users)
    
    conn.commit()
    conn.close()

# Initialize database on startup
init_db()

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/dashboard')
def dashboard():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) FROM assets')
    total_assets = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM workflows WHERE status != "Completed"')
    active_workflows = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(DISTINCT region) FROM assets')
    regions = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM users')
    total_users = cursor.fetchone()[0]
    
    conn.close()
    
    return jsonify({
        'total_assets': total_assets,
        'active_workflows': active_workflows,
        'regions': regions,
        'total_users': total_users
    })

@app.route('/api/assets', methods=['GET', 'POST'])
def handle_assets():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute('SELECT * FROM assets ORDER BY created_at DESC')
        assets = cursor.fetchall()
        conn.close()
        
        # Get column names
        cursor = sqlite3.connect(DATABASE_PATH).cursor()
        cursor.execute('PRAGMA table_info(assets)')
        columns = [column[1] for column in cursor.fetchall()]
        cursor.close()
        
        assets_list = []
        for asset in assets:
            asset_dict = dict(zip(columns, asset))
            assets_list.append(asset_dict)
        
        return jsonify(assets_list)
    
    elif request.method == 'POST':
        data = request.json
        
        cursor.execute('SELECT COUNT(*) FROM assets')
        count = cursor.fetchone()[0] + 1
        asset_id = f'AST-{count:03d}'
        
        try:
            # Insert all 58 MOE fields
            cursor.execute('''
                INSERT INTO assets (
                    asset_id, asset_name, asset_type, asset_status, asset_category, priority_level, strategic_importance,
                    planning_status, need_assessment, feasibility_study, approval_status,
                    location_score, accessibility_rating, infrastructure_quality,
                    investment_proposal, identified_obstacles, mitigation_strategies,
                    financial_obligations, loan_covenants, payment_schedule,
                    electricity_connection, water_connection, sewage_connection, telecommunications,
                    ownership_type, owner_name, ownership_percentage, ownership_documents,
                    land_area, zoning_classification, land_use_permit,
                    total_built_area, usable_area, common_area, parking_area, green_space_area,
                    construction_status, completion_percentage, construction_start_date, expected_completion_date,
                    length_meters, width_meters, height_meters, floor_count,
                    north_boundary, south_boundary, east_boundary, west_boundary,
                    boundary_length_north, boundary_length_south, boundary_length_east, boundary_length_west,
                    region, city, district, street_address, latitude, longitude, elevation_meters,
                    investment_value, maintenance_cost
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                asset_id,
                data.get('asset_name'), data.get('asset_type'), data.get('asset_status'), data.get('asset_category'), data.get('priority_level'), data.get('strategic_importance'),
                data.get('planning_status'), data.get('need_assessment'), data.get('feasibility_study'), data.get('approval_status'),
                data.get('location_score'), data.get('accessibility_rating'), data.get('infrastructure_quality'),
                data.get('investment_proposal'), data.get('identified_obstacles'), data.get('mitigation_strategies'),
                data.get('financial_obligations'), data.get('loan_covenants'), data.get('payment_schedule'),
                data.get('electricity_connection'), data.get('water_connection'), data.get('sewage_connection'), data.get('telecommunications'),
                data.get('ownership_type'), data.get('owner_name'), data.get('ownership_percentage'), data.get('ownership_documents'),
                data.get('land_area'), data.get('zoning_classification'), data.get('land_use_permit'),
                data.get('total_built_area'), data.get('usable_area'), data.get('common_area'), data.get('parking_area'), data.get('green_space_area'),
                data.get('construction_status'), data.get('completion_percentage'), data.get('construction_start_date'), data.get('expected_completion_date'),
                data.get('length_meters'), data.get('width_meters'), data.get('height_meters'), data.get('floor_count'),
                data.get('north_boundary'), data.get('south_boundary'), data.get('east_boundary'), data.get('west_boundary'),
                data.get('boundary_length_north'), data.get('boundary_length_south'), data.get('boundary_length_east'), data.get('boundary_length_west'),
                data.get('region'), data.get('city'), data.get('district'), data.get('street_address'), data.get('latitude'), data.get('longitude'), data.get('elevation_meters'),
                data.get('investment_value'), data.get('maintenance_cost')
            ))
            
            asset_db_id = cursor.lastrowid
            
            # Process uploaded files if any
            uploaded_files = data.get('uploaded_files', [])
            for file_info in uploaded_files:
                cursor.execute('''
                    INSERT INTO files (asset_id, file_name, file_path, file_type, document_category, file_size, ocr_text)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    asset_db_id,
                    file_info['name'],
                    file_info['path'],
                    file_info['type'],
                    file_info['category'],
                    file_info['size'],
                    file_info['ocr_text']
                ))
            
            conn.commit()
            conn.close()
            
            return jsonify({
                'success': True, 
                'asset_id': asset_id, 
                'message': f'Asset {asset_id} created successfully with {len(uploaded_files)} documents processed'
            })
            
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/assets/<int:asset_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_single_asset(asset_id):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute('SELECT * FROM assets WHERE id = ?', (asset_id,))
        asset = cursor.fetchone()
        
        if asset:
            # Get column names
            cursor.execute('PRAGMA table_info(assets)')
            columns = [column[1] for column in cursor.fetchall()]
            asset_dict = dict(zip(columns, asset))
            
            # Get associated files
            cursor.execute('SELECT * FROM files WHERE asset_id = ?', (asset_id,))
            files = cursor.fetchall()
            asset_dict['files'] = files
            
            conn.close()
            return jsonify(asset_dict)
        else:
            conn.close()
            return jsonify({'error': 'Asset not found'}), 404
    
    elif request.method == 'PUT':
        try:
            data = request.json
            
            # Update all fields
            cursor.execute('''
                UPDATE assets SET 
                    asset_name = ?, asset_type = ?, asset_status = ?, asset_category = ?, priority_level = ?, strategic_importance = ?,
                    planning_status = ?, need_assessment = ?, feasibility_study = ?, approval_status = ?,
                    location_score = ?, accessibility_rating = ?, infrastructure_quality = ?,
                    investment_proposal = ?, identified_obstacles = ?, mitigation_strategies = ?,
                    financial_obligations = ?, loan_covenants = ?, payment_schedule = ?,
                    electricity_connection = ?, water_connection = ?, sewage_connection = ?, telecommunications = ?,
                    ownership_type = ?, owner_name = ?, ownership_percentage = ?, ownership_documents = ?,
                    land_area = ?, zoning_classification = ?, land_use_permit = ?,
                    total_built_area = ?, usable_area = ?, common_area = ?, parking_area = ?, green_space_area = ?,
                    construction_status = ?, completion_percentage = ?, construction_start_date = ?, expected_completion_date = ?,
                    length_meters = ?, width_meters = ?, height_meters = ?, floor_count = ?,
                    north_boundary = ?, south_boundary = ?, east_boundary = ?, west_boundary = ?,
                    boundary_length_north = ?, boundary_length_south = ?, boundary_length_east = ?, boundary_length_west = ?,
                    region = ?, city = ?, district = ?, street_address = ?, latitude = ?, longitude = ?, elevation_meters = ?,
                    investment_value = ?, maintenance_cost = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (
                data.get('asset_name'), data.get('asset_type'), data.get('asset_status'), data.get('asset_category'), data.get('priority_level'), data.get('strategic_importance'),
                data.get('planning_status'), data.get('need_assessment'), data.get('feasibility_study'), data.get('approval_status'),
                data.get('location_score'), data.get('accessibility_rating'), data.get('infrastructure_quality'),
                data.get('investment_proposal'), data.get('identified_obstacles'), data.get('mitigation_strategies'),
                data.get('financial_obligations'), data.get('loan_covenants'), data.get('payment_schedule'),
                data.get('electricity_connection'), data.get('water_connection'), data.get('sewage_connection'), data.get('telecommunications'),
                data.get('ownership_type'), data.get('owner_name'), data.get('ownership_percentage'), data.get('ownership_documents'),
                data.get('land_area'), data.get('zoning_classification'), data.get('land_use_permit'),
                data.get('total_built_area'), data.get('usable_area'), data.get('common_area'), data.get('parking_area'), data.get('green_space_area'),
                data.get('construction_status'), data.get('completion_percentage'), data.get('construction_start_date'), data.get('expected_completion_date'),
                data.get('length_meters'), data.get('width_meters'), data.get('height_meters'), data.get('floor_count'),
                data.get('north_boundary'), data.get('south_boundary'), data.get('east_boundary'), data.get('west_boundary'),
                data.get('boundary_length_north'), data.get('boundary_length_south'), data.get('boundary_length_east'), data.get('boundary_length_west'),
                data.get('region'), data.get('city'), data.get('district'), data.get('street_address'), data.get('latitude'), data.get('longitude'), data.get('elevation_meters'),
                data.get('investment_value'), data.get('maintenance_cost'), asset_id
            ))
            
            conn.commit()
            
            if cursor.rowcount > 0:
                conn.close()
                return jsonify({'success': True, 'message': 'Asset updated successfully'})
            else:
                conn.close()
                return jsonify({'error': 'Asset not found'}), 404
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        # Delete associated files first
        cursor.execute('SELECT file_path FROM files WHERE asset_id = ?', (asset_id,))
        files = cursor.fetchall()
        for file_path, in files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except:
                pass
        
        cursor.execute('DELETE FROM files WHERE asset_id = ?', (asset_id,))
        cursor.execute('DELETE FROM assets WHERE id = ?', (asset_id,))
        conn.commit()
        
        if cursor.rowcount > 0:
            conn.close()
            return jsonify({'success': True, 'message': 'Asset deleted successfully'})
        else:
            conn.close()
            return jsonify({'error': 'Asset not found'}), 404

@app.route('/api/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        document_type = request.form.get('document_type', 'general')
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if file and allowed_file(file.filename):
            # Generate unique filename
            filename = secure_filename(file.filename)
            unique_filename = f"{uuid.uuid4()}_{filename}"
            
            # Save file to appropriate directory
            file_path = os.path.join(UPLOAD_FOLDER, document_type, unique_filename)
            file.save(file_path)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            # Extract text using OCR
            ocr_text = extract_text_from_file(file_path)
            
            return jsonify({
                'success': True,
                'file_info': {
                    'name': filename,
                    'path': file_path,
                    'type': file.content_type,
                    'category': document_type,
                    'size': file_size,
                    'ocr_text': ocr_text[:1000] if ocr_text else '',  # Limit OCR text length
                    'processing_details': {
                        'file_size_mb': round(file_size / (1024 * 1024), 2),
                        'text_length': len(ocr_text) if ocr_text else 0,
                        'processing_status': 'success' if ocr_text else 'no_text_found'
                    }
                }
            })
        else:
            return jsonify({'success': False, 'error': 'File type not allowed'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Workflow endpoints (same as before but with permanent storage)
@app.route('/api/workflows', methods=['GET', 'POST'])
def handle_workflows():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute('SELECT * FROM workflows ORDER BY created_at DESC')
        workflows = cursor.fetchall()
        conn.close()
        
        workflows_list = []
        for workflow in workflows:
            workflows_list.append({
                'id': workflow[0],
                'workflow_id': workflow[1],
                'title': workflow[2],
                'description': workflow[3],
                'priority': workflow[4],
                'status': workflow[5],
                'assigned_to': workflow[6],
                'due_date': workflow[7],
                'created_at': workflow[8],
                'updated_at': workflow[9]
            })
        
        return jsonify(workflows_list)
    
    elif request.method == 'POST':
        data = request.json
        
        cursor.execute('SELECT COUNT(*) FROM workflows')
        count = cursor.fetchone()[0] + 1
        workflow_id = f'WF-{count:03d}'
        
        try:
            cursor.execute('''
                INSERT INTO workflows (workflow_id, title, description, priority, status, assigned_to, due_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                workflow_id,
                data.get('title'),
                data.get('description'),
                data.get('priority'),
                data.get('status'),
                data.get('assigned_to'),
                data.get('due_date')
            ))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'workflow_id': workflow_id, 'message': f'Workflow {workflow_id} created successfully'})
            
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workflows/<int:workflow_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_single_workflow(workflow_id):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute('SELECT * FROM workflows WHERE id = ?', (workflow_id,))
        workflow = cursor.fetchone()
        conn.close()
        
        if workflow:
            return jsonify({
                'id': workflow[0],
                'workflow_id': workflow[1],
                'title': workflow[2],
                'description': workflow[3],
                'priority': workflow[4],
                'status': workflow[5],
                'assigned_to': workflow[6],
                'due_date': workflow[7],
                'created_at': workflow[8],
                'updated_at': workflow[9]
            })
        else:
            return jsonify({'error': 'Workflow not found'}), 404
    
    elif request.method == 'PUT':
        try:
            data = request.json
            cursor.execute('''
                UPDATE workflows SET title = ?, description = ?, priority = ?, status = ?, assigned_to = ?, due_date = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (
                data['title'], data['description'], data['priority'], 
                data['status'], data['assigned_to'], data['due_date'], workflow_id
            ))
            conn.commit()
            
            if cursor.rowcount > 0:
                conn.close()
                return jsonify({'success': True, 'message': 'Workflow updated successfully'})
            else:
                conn.close()
                return jsonify({'error': 'Workflow not found'}), 404
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        cursor.execute('DELETE FROM workflows WHERE id = ?', (workflow_id,))
        conn.commit()
        
        if cursor.rowcount > 0:
            conn.close()
            return jsonify({'success': True, 'message': 'Workflow deleted successfully'})
        else:
            conn.close()
            return jsonify({'error': 'Workflow not found'}), 404

# User endpoints (same as before but with permanent storage)
@app.route('/api/users', methods=['GET', 'POST'])
def handle_users():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute('SELECT * FROM users ORDER BY created_date DESC')
        users = cursor.fetchall()
        conn.close()
        
        users_list = []
        for user in users:
            users_list.append({
                'id': user[0],
                'user_id': user[1],
                'full_name': user[2],
                'email': user[3],
                'department': user[4],
                'role': user[5],
                'region': user[6],
                'phone': user[7],
                'created_date': user[8]
            })
        
        return jsonify(users_list)
    
    elif request.method == 'POST':
        data = request.json
        
        cursor.execute('SELECT COUNT(*) FROM users')
        count = cursor.fetchone()[0] + 1
        user_id = f'USR-{count:03d}'
        
        try:
            cursor.execute('''
                INSERT INTO users (user_id, full_name, email, department, role, region, phone)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                user_id,
                data.get('full_name'),
                data.get('email'),
                data.get('department'),
                data.get('role'),
                data.get('region'),
                data.get('phone')
            ))
            
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'user_id': user_id, 'message': f'User {user_id} created successfully'})
            
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_single_user(user_id):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    if request.method == 'GET':
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            return jsonify({
                'id': user[0],
                'user_id': user[1],
                'full_name': user[2],
                'email': user[3],
                'department': user[4],
                'role': user[5],
                'region': user[6],
                'phone': user[7],
                'created_date': user[8]
            })
        else:
            return jsonify({'error': 'User not found'}), 404
    
    elif request.method == 'PUT':
        try:
            data = request.json
            cursor.execute('''
                UPDATE users SET full_name = ?, email = ?, department = ?, role = ?, region = ?, phone = ?
                WHERE id = ?
            ''', (
                data['full_name'], data['email'], data['department'], 
                data['role'], data['region'], data.get('phone', ''), user_id
            ))
            conn.commit()
            
            if cursor.rowcount > 0:
                conn.close()
                return jsonify({'success': True, 'message': 'User updated successfully'})
            else:
                conn.close()
                return jsonify({'error': 'User not found'}), 404
        except Exception as e:
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        
        if cursor.rowcount > 0:
            conn.close()
            return jsonify({'success': True, 'message': 'User deleted successfully'})
        else:
            conn.close()
            return jsonify({'error': 'User not found'}), 404

@app.route('/api/reports/<report_type>')
def generate_report(report_type):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    if report_type == 'assets':
        cursor.execute('SELECT * FROM assets')
        columns = [description[0] for description in cursor.description]
        writer.writerow(columns)
        writer.writerows(cursor.fetchall())
        filename = 'assets_complete_report.csv'
        
    elif report_type == 'regional':
        cursor.execute('SELECT region, COUNT(*) as asset_count, AVG(investment_value) as avg_investment, SUM(investment_value) as total_investment FROM assets GROUP BY region')
        writer.writerow(['Region', 'Asset Count', 'Average Investment', 'Total Investment'])
        writer.writerows(cursor.fetchall())
        filename = 'regional_analysis_report.csv'
        
    elif report_type == 'construction':
        cursor.execute('SELECT construction_status, COUNT(*) as count, AVG(completion_percentage) as avg_completion, SUM(investment_value) as total_investment FROM assets GROUP BY construction_status')
        writer.writerow(['Construction Status', 'Count', 'Average Completion %', 'Total Investment'])
        writer.writerows(cursor.fetchall())
        filename = 'construction_status_report.csv'
        
    elif report_type == 'financial':
        cursor.execute('SELECT asset_type, COUNT(*) as count, SUM(investment_value) as total_investment, AVG(maintenance_cost) as avg_maintenance, SUM(maintenance_cost) as total_maintenance FROM assets GROUP BY asset_type')
        writer.writerow(['Asset Type', 'Count', 'Total Investment', 'Average Maintenance Cost', 'Total Maintenance Cost'])
        writer.writerows(cursor.fetchall())
        filename = 'financial_analysis_report.csv'
    
    else:
        conn.close()
        return jsonify({'error': 'Invalid report type'}), 400
    
    conn.close()
    
    output.seek(0)
    return output.getvalue(), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename={filename}'
    }

# HTML Template with all 58 MOE fields
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Madares Business - Asset Management System</title>
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f5f5; }
        
        .header { background: linear-gradient(135deg, #e67e22, #f39c12); color: white; padding: 1.5rem; text-align: center; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header h1 { font-size: 1.8rem; font-weight: 600; }
        
        .login-container { max-width: 400px; margin: 5rem auto; padding: 2rem; background: white; border-radius: 10px; box-shadow: 0 5px 20px rgba(0,0,0,0.1); }
        .login-form { display: flex; flex-direction: column; gap: 1rem; }
        
        .nav-tabs { background: white; border-bottom: 1px solid #ddd; }
        .nav-tabs ul { display: flex; list-style: none; padding: 0 2rem; }
        .nav-tabs li { margin-right: 2rem; }
        .nav-tabs a { display: block; padding: 1rem 0; text-decoration: none; color: #666; border-bottom: 3px solid transparent; transition: all 0.3s ease; }
        .nav-tabs a.active, .nav-tabs a:hover { color: #e67e22; border-bottom-color: #e67e22; }
        
        .content { padding: 2rem; }
        .dashboard { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 1.5rem; margin-bottom: 2rem; }
        .stat-card { background: white; padding: 1.5rem; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); text-align: center; }
        .stat-card h3 { color: #e67e22; font-size: 2rem; margin-bottom: 0.5rem; }
        .stat-card p { color: #666; font-size: 0.9rem; }
        
        .table-container { background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.05); overflow: hidden; margin-bottom: 2rem; }
        .table-header { padding: 1.5rem; border-bottom: 1px solid #eee; display: flex; justify-content: space-between; align-items: center; }
        .table-header h2 { color: #333; font-size: 1.3rem; }
        
        table { width: 100%; border-collapse: collapse; }
        th, td { padding: 1rem; text-align: left; border-bottom: 1px solid #eee; }
        th { background: #f8f9fa; font-weight: 600; color: #333; }
        tr:hover { background: #f8f9fa; }
        
        .btn { background: #e67e22; color: white; border: none; padding: 0.75rem 1.5rem; border-radius: 5px; cursor: pointer; text-decoration: none; display: inline-block; transition: background 0.3s ease; }
        .btn:hover { background: #d35400; }
        .btn-small { padding: 0.5rem 1rem; font-size: 0.85rem; }
        .btn-secondary { background: #95a5a6; }
        .btn-secondary:hover { background: #7f8c8d; }
        .btn-danger { background: #e74c3c; }
        .btn-danger:hover { background: #c0392b; }
        
        .form-group { margin-bottom: 1rem; }
        .form-group label { display: block; margin-bottom: 0.5rem; color: #333; font-weight: 500; }
        .form-group input, .form-group select, .form-group textarea { width: 100%; padding: 0.75rem; border: 1px solid #ddd; border-radius: 5px; font-size: 1rem; }
        .form-group input:focus, .form-group select:focus, .form-group textarea:focus { outline: none; border-color: #e67e22; box-shadow: 0 0 0 2px rgba(230, 126, 34, 0.1); }
        
        .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 1.5rem; }
        .form-section { background: #f8f9fa; padding: 1.5rem; border-radius: 8px; margin-bottom: 1rem; }
        .form-section h3 { margin-bottom: 1rem; color: #333; cursor: pointer; display: flex; justify-content: space-between; align-items: center; }
        .form-section.collapsed .form-fields { display: none; }
        .form-fields { display: grid; gap: 1rem; }
        
        .upload-area { border: 2px dashed #ddd; border-radius: 8px; padding: 2rem; text-align: center; cursor: pointer; transition: all 0.3s ease; margin: 1rem 0; }
        .upload-area:hover { border-color: #e67e22; background: #fef9f5; }
        .upload-area.dragover { border-color: #e67e22; background: #fef9f5; }
        .upload-area input[type="file"] { display: none; }
        .upload-status { margin-top: 0.5rem; font-size: 0.9rem; }
        .upload-success { color: #27ae60; }
        .upload-error { color: #e74c3c; }
        .upload-processing { color: #f39c12; }
        
        .upload-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 1rem; margin-top: 1rem; }
        .upload-area { border: 2px dashed #ddd; border-radius: 8px; padding: 1.5rem 1rem; text-align: center; cursor: pointer; transition: all 0.3s ease; background: #fafafa; margin: 0; }
        .upload-area:hover { border-color: #d4a574; background: #f9f7f4; }
        .upload-area.uploading { border-color: #d4a574; background: #fff8f0; }
        .upload-area.uploaded { border-color: #28a745; background: #f8fff9; }
        .upload-icon { font-size: 2rem; margin-bottom: 0.5rem; }
        .upload-text { font-weight: 600; color: #333; margin-bottom: 0.5rem; }
        .upload-status { font-size: 0.9rem; color: #666; margin-top: 0.5rem; }
        .upload-status.processing { color: #d4a574; }
        .upload-status.completed { color: #28a745; }
        
        .map-container { height: 300px; border: 1px solid #ddd; border-radius: 5px; margin: 1rem 0; }
        
        .modal { display: none; position: fixed; z-index: 1000; left: 0; top: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.5); }
        .modal-content { background: white; margin: 2% auto; padding: 2rem; width: 95%; max-width: 1200px; border-radius: 10px; max-height: 90vh; overflow-y: auto; }
        .modal-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 1.5rem; }
        .close { font-size: 2rem; cursor: pointer; color: #999; }
        .close:hover { color: #333; }
        
        .alert { padding: 1rem; margin: 1rem 0; border-radius: 5px; }
        .alert-success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
        .alert-error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
        
        .tab-content { display: none; }
        .tab-content.active { display: block; }
        
        .status-badge { padding: 0.25rem 0.75rem; border-radius: 15px; font-size: 0.8rem; font-weight: 500; }
        .status-planning { background: #fff3cd; color: #856404; }
        .status-progress { background: #cce5ff; color: #004085; }
        .status-completed { background: #d4edda; color: #155724; }
        .status-hold { background: #f8d7da; color: #721c24; }
        .status-active { background: #d4edda; color: #155724; }
        
        .priority-high { color: #e74c3c; font-weight: bold; }
        .priority-medium { color: #f39c12; font-weight: bold; }
        .priority-low { color: #27ae60; font-weight: bold; }
        
        @media (max-width: 768px) {
            .header { padding: 1rem; }
            .header h1 { font-size: 1.4rem; }
            .nav-tabs ul { flex-wrap: wrap; gap: 1rem; }
            .content { padding: 1rem; }
            .dashboard { grid-template-columns: 1fr; }
            .form-grid { grid-template-columns: 1fr; }
            .modal-content { width: 98%; margin: 1% auto; padding: 1rem; }
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>Madares Business - Asset Management System</h1>
    </div>

    <!-- Login Form -->
    <div id="loginContainer" class="login-container">
        <h2 style="text-align: center; margin-bottom: 2rem; color: #333;">System Login</h2>
        <form id="loginForm" class="login-form">
            <div class="form-group">
                <label>Username</label>
                <input type="text" id="username" required>
            </div>
            <div class="form-group">
                <label>Password</label>
                <input type="password" id="password" required>
            </div>
            <button type="submit" class="btn">Sign In</button>
        </form>
    </div>

    <!-- Main Application -->
    <div id="mainApp" style="display: none;">
        <nav class="nav-tabs">
            <ul>
                <li><a href="#" onclick="showTab('dashboard')" class="active">Dashboard</a></li>
                <li><a href="#" onclick="showTab('assets')">Assets</a></li>
                <li><a href="#" onclick="showTab('add-asset')">Add Asset</a></li>
                <li><a href="#" onclick="showTab('workflows')">Workflows</a></li>
                <li><a href="#" onclick="showTab('users')">Users</a></li>
                <li><a href="#" onclick="showTab('reports')">Reports</a></li>
            </ul>
        </nav>

        <div class="content">
            <!-- Dashboard Tab -->
            <div id="dashboard" class="tab-content active">
                <div class="dashboard">
                    <div class="stat-card">
                        <h3 id="totalAssets">0</h3>
                        <p>Total Assets</p>
                    </div>
                    <div class="stat-card">
                        <h3 id="activeWorkflows">0</h3>
                        <p>Active Workflows</p>
                    </div>
                    <div class="stat-card">
                        <h3 id="regions">0</h3>
                        <p>Regions</p>
                    </div>
                    <div class="stat-card">
                        <h3 id="totalUsers">0</h3>
                        <p>Total Users</p>
                    </div>
                </div>
                
                <div class="table-container">
                    <div class="table-header">
                        <h2>System Overview</h2>
                    </div>
                    <div style="padding: 2rem; text-align: center; color: #666;">
                        <p>Welcome to Madares Business Asset Management System!</p>
                        <p style="margin-top: 0.5rem;">Complete system with all 58 MOE fields, file upload, and OCR processing.</p>
                        <p style="margin-top: 0.5rem;">Use the navigation tabs above to manage your assets, workflows, and users.</p>
                    </div>
                </div>
            </div>

            <!-- Assets Tab -->
            <div id="assets" class="tab-content">
                <div class="table-container">
                    <div class="table-header">
                        <h2>Asset Management</h2>
                        <input type="text" id="assetSearch" placeholder="Search assets..." style="padding: 0.5rem; border: 1px solid #ddd; border-radius: 5px; width: 250px;">
                    </div>
                    <table>
                        <thead>
                            <tr>
                                <th>Asset ID</th>
                                <th>Name</th>
                                <th>Type</th>
                                <th>Region</th>
                                <th>Status</th>
                                <th>Investment (SAR)</th>
                                <th>Completion</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody id="assetsTableBody">
                            <!-- Assets will be loaded here -->
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- Add Asset Tab with all 58 MOE fields -->
            <div id="add-asset" class="tab-content">
                <div class="table-container">
                    <div class="table-header">
                        <h2>Add New Asset - Complete MOE Form</h2>
                    </div>
                    <div style="padding: 2rem;">
                        <form id="assetForm">
                            <!-- Section 1: Asset Identification & Status -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">1. Asset Identification & Status <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Asset Name *</label>
                                        <input type="text" name="asset_name" required>
                                    </div>
                                    <div class="form-group">
                                        <label>Asset Type *</label>
                                        <select name="asset_type" required>
                                            <option value="">Select Type</option>
                                            <option value="Educational">Educational</option>
                                            <option value="Commercial">Commercial</option>
                                            <option value="Industrial">Industrial</option>
                                            <option value="Residential">Residential</option>
                                            <option value="Healthcare">Healthcare</option>
                                            <option value="Government">Government</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Asset Status</label>
                                        <select name="asset_status">
                                            <option value="Active">Active</option>
                                            <option value="Inactive">Inactive</option>
                                            <option value="Under Review">Under Review</option>
                                            <option value="Disposed">Disposed</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Asset Category</label>
                                        <select name="asset_category">
                                            <option value="Primary">Primary</option>
                                            <option value="Secondary">Secondary</option>
                                            <option value="Support">Support</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Priority Level</label>
                                        <select name="priority_level">
                                            <option value="High">High</option>
                                            <option value="Medium">Medium</option>
                                            <option value="Low">Low</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Strategic Importance</label>
                                        <select name="strategic_importance">
                                            <option value="Critical">Critical</option>
                                            <option value="Important">Important</option>
                                            <option value="Standard">Standard</option>
                                        </select>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 2: Planning & Need Assessment -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">2. Planning & Need Assessment <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Planning Status</label>
                                        <select name="planning_status">
                                            <option value="Not Started">Not Started</option>
                                            <option value="In Progress">In Progress</option>
                                            <option value="Approved">Approved</option>
                                            <option value="Rejected">Rejected</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Need Assessment</label>
                                        <textarea name="need_assessment" rows="3" placeholder="Describe the need assessment results"></textarea>
                                    </div>
                                    <div class="form-group">
                                        <label>Feasibility Study</label>
                                        <select name="feasibility_study">
                                            <option value="Not Required">Not Required</option>
                                            <option value="Required">Required</option>
                                            <option value="In Progress">In Progress</option>
                                            <option value="Completed">Completed</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Approval Status</label>
                                        <select name="approval_status">
                                            <option value="Pending">Pending</option>
                                            <option value="Approved">Approved</option>
                                            <option value="Conditional">Conditional</option>
                                            <option value="Rejected">Rejected</option>
                                        </select>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 3: Location Attractiveness -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">3. Location Attractiveness <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Location Score (1-100)</label>
                                        <input type="number" name="location_score" min="1" max="100">
                                    </div>
                                    <div class="form-group">
                                        <label>Accessibility Rating</label>
                                        <select name="accessibility_rating">
                                            <option value="Excellent">Excellent</option>
                                            <option value="Good">Good</option>
                                            <option value="Average">Average</option>
                                            <option value="Poor">Poor</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Infrastructure Quality</label>
                                        <select name="infrastructure_quality">
                                            <option value="High Quality">High Quality</option>
                                            <option value="Standard">Standard</option>
                                            <option value="Basic">Basic</option>
                                            <option value="Needs Improvement">Needs Improvement</option>
                                        </select>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 4: Investment Proposal & Obstacles -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">4. Investment Proposal & Obstacles <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Investment Proposal</label>
                                        <textarea name="investment_proposal" rows="3" placeholder="Describe the investment proposal"></textarea>
                                    </div>
                                    <div class="form-group">
                                        <label>Identified Obstacles</label>
                                        <textarea name="identified_obstacles" rows="3" placeholder="List any identified obstacles"></textarea>
                                    </div>
                                    <div class="form-group">
                                        <label>Mitigation Strategies</label>
                                        <textarea name="mitigation_strategies" rows="3" placeholder="Describe mitigation strategies"></textarea>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 5: Financial Obligations & Covenants -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">5. Financial Obligations & Covenants <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Financial Obligations</label>
                                        <textarea name="financial_obligations" rows="3" placeholder="Describe financial obligations"></textarea>
                                    </div>
                                    <div class="form-group">
                                        <label>Loan Covenants</label>
                                        <textarea name="loan_covenants" rows="3" placeholder="Describe loan covenants if applicable"></textarea>
                                    </div>
                                    <div class="form-group">
                                        <label>Payment Schedule</label>
                                        <textarea name="payment_schedule" rows="3" placeholder="Describe payment schedule"></textarea>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 6: Utilities Information -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">6. Utilities Information <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Electricity Connection</label>
                                        <select name="electricity_connection">
                                            <option value="Connected">Connected</option>
                                            <option value="Not Connected">Not Connected</option>
                                            <option value="Pending">Pending</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Water Connection</label>
                                        <select name="water_connection">
                                            <option value="Connected">Connected</option>
                                            <option value="Not Connected">Not Connected</option>
                                            <option value="Pending">Pending</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Sewage Connection</label>
                                        <select name="sewage_connection">
                                            <option value="Connected">Connected</option>
                                            <option value="Not Connected">Not Connected</option>
                                            <option value="Pending">Pending</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Telecommunications</label>
                                        <select name="telecommunications">
                                            <option value="Fiber Optic">Fiber Optic</option>
                                            <option value="DSL">DSL</option>
                                            <option value="Wireless">Wireless</option>
                                            <option value="Not Available">Not Available</option>
                                        </select>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 7: Ownership Information -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">7. Ownership Information <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Ownership Type</label>
                                        <select name="ownership_type">
                                            <option value="Government">Government</option>
                                            <option value="Private">Private</option>
                                            <option value="Joint Venture">Joint Venture</option>
                                            <option value="Public-Private Partnership">Public-Private Partnership</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Owner Name</label>
                                        <input type="text" name="owner_name">
                                    </div>
                                    <div class="form-group">
                                        <label>Ownership Percentage (%)</label>
                                        <input type="number" name="ownership_percentage" min="0" max="100" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Ownership Documents Status</label>
                                        <select name="ownership_documents">
                                            <option value="Complete">Complete</option>
                                            <option value="Incomplete">Incomplete</option>
                                            <option value="Under Review">Under Review</option>
                                            <option value="Missing">Missing</option>
                                        </select>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 8: Land & Plan Details -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">8. Land & Plan Details <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Land Area (sqm)</label>
                                        <input type="number" name="land_area" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Zoning Classification</label>
                                        <select name="zoning_classification">
                                            <option value="Residential">Residential</option>
                                            <option value="Commercial">Commercial</option>
                                            <option value="Industrial">Industrial</option>
                                            <option value="Mixed Use">Mixed Use</option>
                                            <option value="Educational">Educational</option>
                                            <option value="Healthcare">Healthcare</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Land Use Permit</label>
                                        <select name="land_use_permit">
                                            <option value="Approved">Approved</option>
                                            <option value="Pending">Pending</option>
                                            <option value="Expired">Expired</option>
                                            <option value="Not Required">Not Required</option>
                                        </select>
                                    </div>
                                </div>
                            </div>

                            <!-- Section 9: Asset Area Details -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">9. Asset Area Details <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Total Built Area (sqm)</label>
                                        <input type="number" name="total_built_area" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Usable Area (sqm)</label>
                                        <input type="number" name="usable_area" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Common Area (sqm)</label>
                                        <input type="number" name="common_area" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Parking Area (sqm)</label>
                                        <input type="number" name="parking_area" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Green Space Area (sqm)</label>
                                        <input type="number" name="green_space_area" step="0.01">
                                    </div>
                                </div>
                            </div>

                            <!-- Section 10: Construction Status -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">10. Construction Status <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Construction Status</label>
                                        <select name="construction_status">
                                            <option value="Planning">Planning</option>
                                            <option value="In Progress">In Progress</option>
                                            <option value="Completed">Completed</option>
                                            <option value="On Hold">On Hold</option>
                                            <option value="Cancelled">Cancelled</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>Completion Percentage (%)</label>
                                        <input type="number" name="completion_percentage" min="0" max="100">
                                    </div>
                                    <div class="form-group">
                                        <label>Construction Start Date</label>
                                        <input type="date" name="construction_start_date">
                                    </div>
                                    <div class="form-group">
                                        <label>Expected Completion Date</label>
                                        <input type="date" name="expected_completion_date">
                                    </div>
                                </div>
                            </div>

                            <!-- Section 11: Physical Dimensions -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">11. Physical Dimensions <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Length (meters)</label>
                                        <input type="number" name="length_meters" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Width (meters)</label>
                                        <input type="number" name="width_meters" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Height (meters)</label>
                                        <input type="number" name="height_meters" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>Floor Count</label>
                                        <input type="number" name="floor_count" min="1">
                                    </div>
                                </div>
                            </div>

                            <!-- Section 12: Boundaries -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">12. Boundaries <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>North Boundary</label>
                                        <input type="text" name="north_boundary" placeholder="e.g., Public road, Private property">
                                    </div>
                                    <div class="form-group">
                                        <label>South Boundary</label>
                                        <input type="text" name="south_boundary" placeholder="e.g., Public road, Private property">
                                    </div>
                                    <div class="form-group">
                                        <label>East Boundary</label>
                                        <input type="text" name="east_boundary" placeholder="e.g., Public road, Private property">
                                    </div>
                                    <div class="form-group">
                                        <label>West Boundary</label>
                                        <input type="text" name="west_boundary" placeholder="e.g., Public road, Private property">
                                    </div>
                                    <div class="form-group">
                                        <label>North Boundary Length (meters)</label>
                                        <input type="number" name="boundary_length_north" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>South Boundary Length (meters)</label>
                                        <input type="number" name="boundary_length_south" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>East Boundary Length (meters)</label>
                                        <input type="number" name="boundary_length_east" step="0.01">
                                    </div>
                                    <div class="form-group">
                                        <label>West Boundary Length (meters)</label>
                                        <input type="number" name="boundary_length_west" step="0.01">
                                    </div>
                                </div>
                            </div>

                            <!-- Section 13: Geographic Location -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">13. Geographic Location <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Region *</label>
                                        <select name="region" required>
                                            <option value="">Select Region</option>
                                            <option value="Riyadh">Riyadh</option>
                                            <option value="Makkah">Makkah</option>
                                            <option value="Eastern Province">Eastern Province</option>
                                            <option value="Asir">Asir</option>
                                            <option value="Qassim">Qassim</option>
                                            <option value="Tabuk">Tabuk</option>
                                            <option value="Hail">Hail</option>
                                            <option value="Northern Borders">Northern Borders</option>
                                            <option value="Jazan">Jazan</option>
                                            <option value="Najran">Najran</option>
                                            <option value="Al Bahah">Al Bahah</option>
                                            <option value="Al Jouf">Al Jouf</option>
                                        </select>
                                    </div>
                                    <div class="form-group">
                                        <label>City *</label>
                                        <input type="text" name="city" required>
                                    </div>
                                    <div class="form-group">
                                        <label>District</label>
                                        <input type="text" name="district">
                                    </div>
                                    <div class="form-group">
                                        <label>Street Address</label>
                                        <input type="text" name="street_address">
                                    </div>
                                    <div class="form-group">
                                        <label>Latitude</label>
                                        <input type="number" name="latitude" step="any" id="latInput">
                                    </div>
                                    <div class="form-group">
                                        <label>Longitude</label>
                                        <input type="number" name="longitude" step="any" id="lngInput">
                                    </div>
                                    <div class="form-group">
                                        <label>Elevation (meters)</label>
                                        <input type="number" name="elevation_meters" step="0.01">
                                    </div>
                                </div>
                                <div id="map" class="map-container"></div>
                            </div>

                            <!-- Section 14: Supporting Documents -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">14. Supporting Documents <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Property Deed</label>
                                        <div class="upload-area" onclick="triggerFileUpload('property_deed')">
                                            <input type="file" id="property_deed" accept=".pdf,.doc,.docx,.jpg,.png" onchange="handleFileUpload(this, 'property_deed')">
                                            <div>📄 Click to upload Property Deed</div>
                                            <div class="upload-status" id="property_deed_status"></div>
                                        </div>
                                    </div>
                                    <div class="form-group">
                                        <label>Ownership Documents</label>
                                        <div class="upload-area" onclick="triggerFileUpload('ownership_documents')">
                                            <input type="file" id="ownership_documents" accept=".pdf,.doc,.docx,.jpg,.png" onchange="handleFileUpload(this, 'ownership_documents')">
                                            <div>📋 Click to upload Ownership Documents</div>
                                            <div class="upload-status" id="ownership_documents_status"></div>
                                        </div>
                                    </div>
                                    <div class="form-group">
                                        <label>Construction Plans</label>
                                        <div class="upload-area" onclick="triggerFileUpload('construction_plans')">
                                            <input type="file" id="construction_plans" accept=".pdf,.dwg,.jpg,.png" onchange="handleFileUpload(this, 'construction_plans')">
                                            <div>🏗️ Click to upload Construction Plans</div>
                                            <div class="upload-status" id="construction_plans_status"></div>
                                        </div>
                                    </div>
                                    <div class="form-group">
                                        <label>Financial Documents</label>
                                        <div class="upload-area" onclick="triggerFileUpload('financial_documents')">
                                            <input type="file" id="financial_documents" accept=".pdf,.xls,.xlsx,.doc,.docx" onchange="handleFileUpload(this, 'financial_documents')">
                                            <div>💰 Click to upload Financial Documents</div>
                                            <div class="upload-status" id="financial_documents_status"></div>
                                        </div>
                                    </div>
                                    <div class="form-group">
                                        <label>Legal Documents</label>
                                        <div class="upload-area" onclick="triggerFileUpload('legal_documents')">
                                            <input type="file" id="legal_documents" accept=".pdf,.doc,.docx" onchange="handleFileUpload(this, 'legal_documents')">
                                            <div>⚖️ Click to upload Legal Documents</div>
                                            <div class="upload-status" id="legal_documents_status"></div>
                                        </div>
                                    </div>
                                    <div class="form-group">
                                        <label>Inspection Reports</label>
                                        <div class="upload-area" onclick="triggerFileUpload('inspection_reports')">
                                            <input type="file" id="inspection_reports" accept=".pdf,.doc,.docx,.jpg,.png" onchange="handleFileUpload(this, 'inspection_reports')">
                                            <div>🔍 Click to upload Inspection Reports</div>
                                            <div class="upload-status" id="inspection_reports_status"></div>
                                        </div>
                                    </div>
                                </div>
                            </div>

                            <!-- Financial Information -->
                            <div class="form-section">
                                <h3 onclick="toggleSection(this)">Financial Information <span>▼</span></h3>
                                <div class="form-fields">
                                    <div class="form-group">
                                        <label>Investment Value (SAR) *</label>
                                        <input type="number" name="investment_value" step="0.01" required>
                                    </div>
                                    <div class="form-group">
                                        <label>Annual Maintenance Cost (SAR)</label>
                                        <input type="number" name="maintenance_cost" step="0.01">
                                    </div>
                                </div>
                            </div>

                            <div style="text-align: center; margin-top: 2rem;">
                                  <!-- Supporting Documents Section -->
                        <div class="form-section">
                            <div class="section-header" onclick="toggleSection('supporting-docs')">
                                <h3>Supporting Documents</h3>
                                <span class="toggle-icon">▼</span>
                            </div>
                            <div class="section-content" id="supporting-docs">
                                <div class="upload-grid">
                                    <div class="upload-area" onclick="document.getElementById('property-deed').click()">
                                        <div class="upload-icon">📄</div>
                                        <div class="upload-text">Property Deed</div>
                                        <div class="upload-status" id="property-deed-status">Click to upload</div>
                                        <input type="file" id="property-deed" name="property-deed" accept=".pdf,.doc,.docx,.jpg,.png" style="display:none" onchange="handleFileUpload(this, 'property-deed-status')">
                                    </div>
                                    
                                    <div class="upload-area" onclick="document.getElementById('ownership-docs').click()">
                                        <div class="upload-icon">📋</div>
                                        <div class="upload-text">Ownership Documents</div>
                                        <div class="upload-status" id="ownership-docs-status">Click to upload</div>
                                        <input type="file" id="ownership-docs" name="ownership-docs" accept=".pdf,.doc,.docx,.jpg,.png" style="display:none" onchange="handleFileUpload(this, 'ownership-docs-status')">
                                    </div>
                                    
                                    <div class="upload-area" onclick="document.getElementById('construction-plans').click()">
                                        <div class="upload-icon">🏗️</div>
                                        <div class="upload-text">Construction Plans</div>
                                        <div class="upload-status" id="construction-plans-status">Click to upload</div>
                                        <input type="file" id="construction-plans" name="construction-plans" accept=".pdf,.dwg,.jpg,.png" style="display:none" onchange="handleFileUpload(this, 'construction-plans-status')">
                                    </div>
                                    
                                    <div class="upload-area" onclick="document.getElementById('financial-docs').click()">
                                        <div class="upload-icon">💰</div>
                                        <div class="upload-text">Financial Documents</div>
                                        <div class="upload-status" id="financial-docs-status">Click to upload</div>
                                        <input type="file" id="financial-docs" name="financial-docs" accept=".pdf,.xls,.xlsx,.doc,.docx" style="display:none" onchange="handleFileUpload(this, 'financial-docs-status')">
                                    </div>
                                    
                                    <div class="upload-area" onclick="document.getElementById('legal-docs').click()">
                                        <div class="upload-icon">⚖️</div>
                                        <div class="upload-text">Legal Documents</div>
                                        <div class="upload-status" id="legal-docs-status">Click to upload</div>
                                        <input type="file" id="legal-docs" name="legal-docs" accept=".pdf,.doc,.docx" style="display:none" onchange="handleFileUpload(this, 'legal-docs-status')">
                                    </div>
                                    
                                    <div class="upload-area" onclick="document.getElementById('inspection-reports').click()">
                                        <div class="upload-icon">🔍</div>
                                        <div class="upload-text">Inspection Reports</div>
                                        <div class="upload-status" id="inspection-reports-status">Click to upload</div>
                                        <input type="file" id="inspection-reports" name="inspection-reports" accept=".pdf,.doc,.docx,.jpg,.png" style="display:none" onchange="handleFileUpload(this, 'inspection-reports-status')">
                                    </div>
                                </div>
                            </div>
                        </div>

                        <button type="submit" class="submit-btn">Submit Asset Registration</button>
                    </form>
                </div>         </form>
                    </div>
                </div>
            </div>

            <!-- Workflows Tab -->
            <div id="workflows" class="tab-content">
                <div class="table-container">
                    <div class="table-header">
                        <h2>Workflow Management</h2>
                        <button class="btn" onclick="openWorkflowModal()">Create New Workflow</button>
                    </div>
                    <table>
                        <thead>
                            <tr>
                                <th>Workflow ID</th>
                                <th>Title</th>
                                <th>Priority</th>
                                <th>Status</th>
                                <th>Assigned To</th>
                                <th>Due Date</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody id="workflowsTableBody">
                            <!-- Workflows will be loaded here -->
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- Users Tab -->
            <div id="users" class="tab-content">
                <div class="table-container">
                    <div class="table-header">
                        <h2>User Management</h2>
                        <button class="btn" onclick="openUserModal()">Add New User</button>
                    </div>
                    <table>
                        <thead>
                            <tr>
                                <th>User ID</th>
                                <th>Name</th>
                                <th>Email</th>
                                <th>Role</th>
                                <th>Department</th>
                                <th>Region</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody id="usersTableBody">
                            <!-- Users will be loaded here -->
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- Reports Tab -->
            <div id="reports" class="tab-content">
                <div class="dashboard">
                    <div class="stat-card" style="cursor: pointer;" onclick="generateReport('assets')">
                        <h3>Complete Asset Report</h3>
                        <div style="color: #666; margin-top: 0.5rem;">All 58 MOE fields with documents</div>
                        <div style="margin-top: 1rem;">
                            <button class="btn btn-small">Generate CSV</button>
                        </div>
                    </div>
                    <div class="stat-card" style="cursor: pointer;" onclick="generateReport('regional')">
                        <h3>Regional Analysis</h3>
                        <div style="color: #666; margin-top: 0.5rem;">Assets by region with investments</div>
                        <div style="margin-top: 1rem;">
                            <button class="btn btn-small">Generate CSV</button>
                        </div>
                    </div>
                    <div class="stat-card" style="cursor: pointer;" onclick="generateReport('construction')">
                        <h3>Construction Progress</h3>
                        <div style="color: #666; margin-top: 0.5rem;">Status and completion tracking</div>
                        <div style="margin-top: 1rem;">
                            <button class="btn btn-small">Generate CSV</button>
                        </div>
                    </div>
                    <div class="stat-card" style="cursor: pointer;" onclick="generateReport('financial')">
                        <h3>Financial Analysis</h3>
                        <div style="color: #666; margin-top: 0.5rem;">Investment and maintenance costs</div>
                        <div style="margin-top: 1rem;">
                            <button class="btn btn-small">Generate CSV</button>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- Modals -->
    <div id="assetModal" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h2 id="assetModalTitle">Asset Details</h2>
                <span class="close" onclick="closeModal('assetModal')">&times;</span>
            </div>
            <div id="assetModalContent">
                <!-- Asset details will be loaded here -->
            </div>
        </div>
    </div>

    <div id="workflowModal" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h2 id="workflowModalTitle">Create New Workflow</h2>
                <span class="close" onclick="closeModal('workflowModal')">&times;</span>
            </div>
            <form id="workflowForm">
                <div class="form-group">
                    <label>Title</label>
                    <input type="text" name="title" required>
                </div>
                <div class="form-group">
                    <label>Description</label>
                    <textarea name="description" rows="3"></textarea>
                </div>
                <div class="form-group">
                    <label>Priority</label>
                    <select name="priority" required>
                        <option value="High">High</option>
                        <option value="Medium">Medium</option>
                        <option value="Low">Low</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Status</label>
                    <select name="status" required>
                        <option value="Not Started">Not Started</option>
                        <option value="In Progress">In Progress</option>
                        <option value="Completed">Completed</option>
                        <option value="On Hold">On Hold</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Assigned To</label>
                    <input type="text" name="assigned_to" required>
                </div>
                <div class="form-group">
                    <label>Due Date</label>
                    <input type="date" name="due_date" required>
                </div>
                <div style="text-align: right; margin-top: 1.5rem;">
                    <button type="button" class="btn btn-secondary" onclick="closeModal('workflowModal')">Cancel</button>
                    <button type="submit" class="btn" style="margin-left: 1rem;">Create Workflow</button>
                </div>
            </form>
        </div>
    </div>

    <div id="userModal" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h2 id="userModalTitle">Add New User</h2>
                <span class="close" onclick="closeModal('userModal')">&times;</span>
            </div>
            <form id="userForm">
                <div class="form-group">
                    <label>Full Name</label>
                    <input type="text" name="full_name" required>
                </div>
                <div class="form-group">
                    <label>Email</label>
                    <input type="email" name="email" required>
                </div>
                <div class="form-group">
                    <label>Department</label>
                    <select name="department" required>
                        <option value="Asset Management">Asset Management</option>
                        <option value="Operations">Operations</option>
                        <option value="Finance">Finance</option>
                        <option value="Legal">Legal</option>
                        <option value="IT">IT</option>
                        <option value="HR">HR</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Role</label>
                    <input type="text" name="role" required>
                </div>
                <div class="form-group">
                    <label>Region</label>
                    <select name="region" required>
                        <option value="Riyadh">Riyadh</option>
                        <option value="Jeddah">Jeddah</option>
                        <option value="Dammam">Dammam</option>
                        <option value="Abha">Abha</option>
                        <option value="Buraidah">Buraidah</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Phone</label>
                    <input type="tel" name="phone">
                </div>
                <div style="text-align: right; margin-top: 1.5rem;">
                    <button type="button" class="btn btn-secondary" onclick="closeModal('userModal')">Cancel</button>
                    <button type="submit" class="btn" style="margin-left: 1rem;">Add User</button>
                </div>
            </form>
        </div>
    </div>

    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <script>
        let map;
        let currentEditId = null;
        let currentEditType = null;
        let uploadedFiles = [];

        // Login functionality
        document.getElementById('loginForm').addEventListener('submit', function(e) {
            e.preventDefault();
            const username = document.getElementById('username').value.trim().toLowerCase();
            const password = document.getElementById('password').value.trim();
            
            if (username === 'admin' && password === 'password123') {
                document.getElementById('loginContainer').style.display = 'none';
                document.getElementById('mainApp').style.display = 'block';
                loadDashboard();
                loadAssets();
                loadWorkflows();
                loadUsers();
                initMap();
            } else {
                alert('Invalid credentials. Use admin/password123');
            }
        });

        // Tab functionality
        function showTab(tabName) {
            document.querySelectorAll('.tab-content').forEach(tab => tab.classList.remove('active'));
            document.querySelectorAll('.nav-tabs a').forEach(link => link.classList.remove('active'));
            
            document.getElementById(tabName).classList.add('active');
            event.target.classList.add('active');
            
            if (tabName === 'add-asset' && map) {
                setTimeout(() => map.invalidateSize(), 100);
            }
        }

        // Toggle form sections
        function toggleSection(element) {
            const section = element.parentElement;
            section.classList.toggle('collapsed');
            const arrow = element.querySelector('span');
            arrow.textContent = section.classList.contains('collapsed') ? '▶' : '▼';
        }

        // File upload functionality
        function triggerFileUpload(documentType) {
            document.getElementById(documentType).click();
        }

        function handleFileUpload(input, documentType) {
            const file = input.files[0];
            if (!file) return;

            const statusElement = document.getElementById(documentType + '_status');
            statusElement.innerHTML = '<span class="upload-processing">⏳ Processing...</span>';

            const formData = new FormData();
            formData.append('file', file);
            formData.append('document_type', documentType);

            fetch('/api/upload', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(result => {
                if (result.success) {
                    const fileInfo = result.file_info;
                    uploadedFiles.push(fileInfo);
                    
                    statusElement.innerHTML = `
                        <span class="upload-success">✓ ${fileInfo.name} - OCR processed successfully</span>
                        <div style="font-size: 0.8rem; color: #666; margin-top: 0.25rem;">
                            Size: ${fileInfo.processing_details.file_size_mb}MB | 
                            Text: ${fileInfo.processing_details.text_length} chars | 
                            Status: ${fileInfo.processing_details.processing_status}
                        </div>
                    `;
                } else {
                    statusElement.innerHTML = `<span class="upload-error">❌ Upload failed: ${result.error}</span>`;
                }
            })
            .catch(error => {
                statusElement.innerHTML = `<span class="upload-error">❌ Upload failed: ${error.message}</span>`;
            });
        }

        // Load dashboard data
        function loadDashboard() {
            fetch('/api/dashboard')
                .then(response => response.json())
                .then(data => {
                    document.getElementById('totalAssets').textContent = data.total_assets;
                    document.getElementById('activeWorkflows').textContent = data.active_workflows;
                    document.getElementById('regions').textContent = data.regions;
                    document.getElementById('totalUsers').textContent = data.total_users;
                });
        }

        // Load assets
        function loadAssets() {
            fetch('/api/assets')
                .then(response => response.json())
                .then(assets => {
                    const tbody = document.getElementById('assetsTableBody');
                    tbody.innerHTML = '';
                    
                    assets.forEach(asset => {
                        const statusClass = asset.construction_status ? 
                            `status-${asset.construction_status.toLowerCase().replace(' ', '')}` : '';
                        
                        const row = document.createElement('tr');
                        row.innerHTML = `
                            <td>${asset.asset_id}</td>
                            <td>${asset.asset_name}</td>
                            <td>${asset.asset_type}</td>
                            <td>${asset.region}</td>
                            <td><span class="status-badge ${statusClass}">${asset.construction_status || 'N/A'}</span></td>
                            <td>${asset.investment_value ? asset.investment_value.toLocaleString() : 'N/A'}</td>
                            <td>${asset.completion_percentage || 0}%</td>
                            <td>
                                <button class="btn btn-small" onclick="viewAsset(${asset.id})">View</button>
                                <button class="btn btn-small btn-secondary" onclick="editAsset(${asset.id})" style="margin-left: 0.5rem;">Edit</button>
                                <button class="btn btn-small btn-danger" onclick="deleteAsset(${asset.id})" style="margin-left: 0.5rem;">Delete</button>
                            </td>
                        `;
                        tbody.appendChild(row);
                    });
                });
        }

        // Asset search functionality
        document.getElementById('assetSearch').addEventListener('input', function(e) {
            const searchTerm = e.target.value.toLowerCase();
            const rows = document.querySelectorAll('#assetsTableBody tr');
            
            rows.forEach(row => {
                const text = row.textContent.toLowerCase();
                row.style.display = text.includes(searchTerm) ? '' : 'none';
            });
        });

        // Asset form submission
        document.getElementById('assetForm').addEventListener('submit', function(e) {
            e.preventDefault();
            
            const formData = new FormData(this);
            const data = Object.fromEntries(formData);
            
            // Add uploaded files to the data
            data.uploaded_files = uploadedFiles;
            
            fetch('/api/assets', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(data)
            })
            .then(response => response.json())
            .then(result => {
                if (result.success) {
                    alert(`Asset created successfully: ${result.asset_id}\\n${result.message}`);
                    this.reset();
                    uploadedFiles = [];
                    
                    // Clear upload status
                    document.querySelectorAll('.upload-status').forEach(status => {
                        status.innerHTML = '';
                    });
                    
                    // Clear map markers
                    if (map) {
                        map.eachLayer(function (layer) {
                            if (layer instanceof L.Marker) {
                                map.removeLayer(layer);
                            }
                        });
                    }
                    
                    loadAssets();
                    loadDashboard();
                } else {
                    alert('Error: ' + result.error);
                }
            });
        });

        // View asset
        function viewAsset(id) {
            fetch(`/api/assets/${id}`)
                .then(response => response.json())
                .then(asset => {
                    document.getElementById('assetModalTitle').textContent = `Asset Details - ${asset.asset_id}`;
                    
                    let filesHtml = '';
                    if (asset.files && asset.files.length > 0) {
                        filesHtml = '<h4>Uploaded Documents:</h4><ul>';
                        asset.files.forEach(file => {
                            filesHtml += `<li>${file[1]} (${file[4]}) - ${file[2]}</li>`;
                        });
                        filesHtml += '</ul>';
                    }
                    
                    document.getElementById('assetModalContent').innerHTML = `
                        <div class="form-grid">
                            <div><strong>Asset ID:</strong> ${asset.asset_id}</div>
                            <div><strong>Name:</strong> ${asset.asset_name}</div>
                            <div><strong>Type:</strong> ${asset.asset_type}</div>
                            <div><strong>Status:</strong> ${asset.asset_status || 'N/A'}</div>
                            <div><strong>Region:</strong> ${asset.region}</div>
                            <div><strong>City:</strong> ${asset.city}</div>
                            <div><strong>Investment Value:</strong> ${asset.investment_value ? asset.investment_value.toLocaleString() + ' SAR' : 'N/A'}</div>
                            <div><strong>Construction Status:</strong> ${asset.construction_status || 'N/A'}</div>
                            <div><strong>Completion:</strong> ${asset.completion_percentage || 0}%</div>
                            <div><strong>Maintenance Cost:</strong> ${asset.maintenance_cost ? asset.maintenance_cost.toLocaleString() + ' SAR' : 'N/A'}</div>
                            <div><strong>Coordinates:</strong> ${asset.latitude && asset.longitude ? `${asset.latitude}, ${asset.longitude}` : 'N/A'}</div>
                            <div><strong>Land Area:</strong> ${asset.land_area ? asset.land_area + ' sqm' : 'N/A'}</div>
                        </div>
                        ${filesHtml}
                        <div style="text-align: right; margin-top: 1.5rem;">
                            <button class="btn btn-secondary" onclick="editAsset(${asset.id})">Edit Asset</button>
                        </div>
                    `;
                    document.getElementById('assetModal').style.display = 'block';
                });
        }

        // Edit asset (simplified for brevity - would include all 58 fields)
        function editAsset(id) {
            fetch(`/api/assets/${id}`)
                .then(response => response.json())
                .then(asset => {
                    currentEditId = id;
                    currentEditType = 'asset';
                    
                    document.getElementById('assetModalTitle').textContent = `Edit Asset - ${asset.asset_id}`;
                    document.getElementById('assetModalContent').innerHTML = `
                        <form id="editAssetForm" class="form-grid">
                            <div class="form-group">
                                <label>Asset Name</label>
                                <input type="text" name="asset_name" value="${asset.asset_name || ''}" required>
                            </div>
                            <div class="form-group">
                                <label>Asset Type</label>
                                <select name="asset_type" required>
                                    <option value="Educational" ${asset.asset_type === 'Educational' ? 'selected' : ''}>Educational</option>
                                    <option value="Commercial" ${asset.asset_type === 'Commercial' ? 'selected' : ''}>Commercial</option>
                                    <option value="Industrial" ${asset.asset_type === 'Industrial' ? 'selected' : ''}>Industrial</option>
                                    <option value="Residential" ${asset.asset_type === 'Residential' ? 'selected' : ''}>Residential</option>
                                    <option value="Healthcare" ${asset.asset_type === 'Healthcare' ? 'selected' : ''}>Healthcare</option>
                                    <option value="Government" ${asset.asset_type === 'Government' ? 'selected' : ''}>Government</option>
                                </select>
                            </div>
                            <div class="form-group">
                                <label>Region</label>
                                <select name="region" required>
                                    <option value="Riyadh" ${asset.region === 'Riyadh' ? 'selected' : ''}>Riyadh</option>
                                    <option value="Makkah" ${asset.region === 'Makkah' ? 'selected' : ''}>Makkah</option>
                                    <option value="Eastern Province" ${asset.region === 'Eastern Province' ? 'selected' : ''}>Eastern Province</option>
                                    <option value="Asir" ${asset.region === 'Asir' ? 'selected' : ''}>Asir</option>
                                    <option value="Qassim" ${asset.region === 'Qassim' ? 'selected' : ''}>Qassim</option>
                                </select>
                            </div>
                            <div class="form-group">
                                <label>City</label>
                                <input type="text" name="city" value="${asset.city || ''}" required>
                            </div>
                            <div class="form-group">
                                <label>Investment Value (SAR)</label>
                                <input type="number" name="investment_value" step="0.01" value="${asset.investment_value || ''}" required>
                            </div>
                            <div class="form-group">
                                <label>Construction Status</label>
                                <select name="construction_status">
                                    <option value="Planning" ${asset.construction_status === 'Planning' ? 'selected' : ''}>Planning</option>
                                    <option value="In Progress" ${asset.construction_status === 'In Progress' ? 'selected' : ''}>In Progress</option>
                                    <option value="Completed" ${asset.construction_status === 'Completed' ? 'selected' : ''}>Completed</option>
                                    <option value="On Hold" ${asset.construction_status === 'On Hold' ? 'selected' : ''}>On Hold</option>
                                </select>
                            </div>
                            <div class="form-group">
                                <label>Completion Percentage</label>
                                <input type="number" name="completion_percentage" min="0" max="100" value="${asset.completion_percentage || ''}">
                            </div>
                            <div class="form-group">
                                <label>Maintenance Cost (SAR)</label>
                                <input type="number" name="maintenance_cost" step="0.01" value="${asset.maintenance_cost || ''}">
                            </div>
                            <div style="grid-column: 1 / -1; text-align: right; margin-top: 1.5rem;">
                                <button type="button" class="btn btn-secondary" onclick="closeModal('assetModal')">Cancel</button>
                                <button type="submit" class="btn" style="margin-left: 1rem;">Save Changes</button>
                            </div>
                        </form>
                    `;
                    document.getElementById('assetModal').style.display = 'block';
                    
                    // Add form submission handler
                    document.getElementById('editAssetForm').addEventListener('submit', function(e) {
                        e.preventDefault();
                        
                        const formData = new FormData(this);
                        const data = Object.fromEntries(formData);
                        
                        fetch(`/api/assets/${currentEditId}`, {
                            method: 'PUT',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify(data)
                        })
                        .then(response => response.json())
                        .then(result => {
                            if (result.success) {
                                alert('Asset updated successfully!');
                                closeModal('assetModal');
                                loadAssets();
                                loadDashboard();
                            } else {
                                alert('Error: ' + result.error);
                            }
                        });
                    });
                });
        }

        // Delete asset
        function deleteAsset(id) {
            if (confirm('Are you sure you want to delete this asset?')) {
                fetch(`/api/assets/${id}`, { method: 'DELETE' })
                    .then(response => response.json())
                    .then(result => {
                        if (result.success) {
                            alert('Asset deleted successfully!');
                            loadAssets();
                            loadDashboard();
                        } else {
                            alert('Error: ' + result.error);
                        }
                    });
            }
        }

        // Load workflows (same as before)
        function loadWorkflows() {
            fetch('/api/workflows')
                .then(response => response.json())
                .then(workflows => {
                    const tbody = document.getElementById('workflowsTableBody');
                    tbody.innerHTML = '';
                    
                    workflows.forEach(workflow => {
                        const priorityClass = `priority-${workflow.priority.toLowerCase()}`;
                        const statusClass = workflow.status ? 
                            `status-${workflow.status.toLowerCase().replace(' ', '')}` : '';
                        
                        const row = document.createElement('tr');
                        row.innerHTML = `
                            <td>${workflow.workflow_id}</td>
                            <td>${workflow.title}</td>
                            <td><span class="${priorityClass}">${workflow.priority}</span></td>
                            <td><span class="status-badge ${statusClass}">${workflow.status}</span></td>
                            <td>${workflow.assigned_to}</td>
                            <td>${workflow.due_date}</td>
                            <td>
                                <button class="btn btn-small" onclick="viewWorkflow(${workflow.id})">View</button>
                                <button class="btn btn-small btn-secondary" onclick="editWorkflow(${workflow.id})" style="margin-left: 0.5rem;">Edit</button>
                                <button class="btn btn-small btn-danger" onclick="deleteWorkflow(${workflow.id})" style="margin-left: 0.5rem;">Delete</button>
                            </td>
                        `;
                        tbody.appendChild(row);
                    });
                });
        }

        // Workflow functions (same as before)
        function openWorkflowModal() {
            currentEditId = null;
            currentEditType = null;
            document.getElementById('workflowModalTitle').textContent = 'Create New Workflow';
            document.getElementById('workflowForm').reset();
            document.getElementById('workflowModal').style.display = 'block';
        }

        function viewWorkflow(id) {
            fetch(`/api/workflows/${id}`)
                .then(response => response.json())
                .then(workflow => {
                    alert(`Workflow Details - ${workflow.workflow_id}\\n\\nTitle: ${workflow.title}\\nDescription: ${workflow.description}\\nPriority: ${workflow.priority}\\nStatus: ${workflow.status}\\nAssigned To: ${workflow.assigned_to}\\nDue Date: ${workflow.due_date}`);
                });
        }

        function editWorkflow(id) {
            fetch(`/api/workflows/${id}`)
                .then(response => response.json())
                .then(workflow => {
                    currentEditId = id;
                    currentEditType = 'workflow';
                    
                    document.getElementById('workflowModalTitle').textContent = `Edit Workflow - ${workflow.workflow_id}`;
                    
                    const form = document.getElementById('workflowForm');
                    form.title.value = workflow.title;
                    form.description.value = workflow.description;
                    form.priority.value = workflow.priority;
                    form.status.value = workflow.status;
                    form.assigned_to.value = workflow.assigned_to;
                    form.due_date.value = workflow.due_date;
                    
                    document.getElementById('workflowModal').style.display = 'block';
                });
        }

        function deleteWorkflow(id) {
            if (confirm('Are you sure you want to delete this workflow?')) {
                fetch(`/api/workflows/${id}`, { method: 'DELETE' })
                    .then(response => response.json())
                    .then(result => {
                        if (result.success) {
                            alert('Workflow deleted successfully!');
                            loadWorkflows();
                            loadDashboard();
                        } else {
                            alert('Error: ' + result.error);
                        }
                    });
            }
        }

        document.getElementById('workflowForm').addEventListener('submit', function(e) {
            e.preventDefault();
            
            const formData = new FormData(this);
            const data = Object.fromEntries(formData);
            
            const method = currentEditId ? 'PUT' : 'POST';
            const url = currentEditId ? `/api/workflows/${currentEditId}` : '/api/workflows';
            
            fetch(url, {
                method: method,
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(data)
            })
            .then(response => response.json())
            .then(result => {
                if (result.success) {
                    alert(currentEditId ? 'Workflow updated successfully!' : 'Workflow created successfully: ' + result.workflow_id);
                    closeModal('workflowModal');
                    loadWorkflows();
                    loadDashboard();
                } else {
                    alert('Error: ' + result.error);
                }
            });
        });

        // Load users (same as before)
        function loadUsers() {
            fetch('/api/users')
                .then(response => response.json())
                .then(users => {
                    const tbody = document.getElementById('usersTableBody');
                    tbody.innerHTML = '';
                    
                    users.forEach(user => {
                        const row = document.createElement('tr');
                        row.innerHTML = `
                            <td>${user.user_id}</td>
                            <td>${user.full_name}</td>
                            <td>${user.email}</td>
                            <td>${user.role}</td>
                            <td>${user.department}</td>
                            <td>${user.region}</td>
                            <td>
                                <button class="btn btn-small" onclick="viewUser(${user.id})">View</button>
                                <button class="btn btn-small btn-secondary" onclick="editUser(${user.id})" style="margin-left: 0.5rem;">Edit</button>
                                <button class="btn btn-small btn-danger" onclick="deleteUser(${user.id})" style="margin-left: 0.5rem;">Delete</button>
                            </td>
                        `;
                        tbody.appendChild(row);
                    });
                });
        }

        // User functions (same as before)
        function openUserModal() {
            currentEditId = null;
            currentEditType = null;
            document.getElementById('userModalTitle').textContent = 'Add New User';
            document.getElementById('userForm').reset();
            document.getElementById('userModal').style.display = 'block';
        }

        function viewUser(id) {
            fetch(`/api/users/${id}`)
                .then(response => response.json())
                .then(user => {
                    alert(`User Details - ${user.user_id}\\n\\nName: ${user.full_name}\\nEmail: ${user.email}\\nDepartment: ${user.department}\\nRole: ${user.role}\\nRegion: ${user.region}\\nPhone: ${user.phone || 'N/A'}`);
                });
        }

        function editUser(id) {
            fetch(`/api/users/${id}`)
                .then(response => response.json())
                .then(user => {
                    currentEditId = id;
                    currentEditType = 'user';
                    
                    document.getElementById('userModalTitle').textContent = `Edit User - ${user.user_id}`;
                    
                    const form = document.getElementById('userForm');
                    form.full_name.value = user.full_name;
                    form.email.value = user.email;
                    form.department.value = user.department;
                    form.role.value = user.role;
                    form.region.value = user.region;
                    form.phone.value = user.phone || '';
                    
                    document.getElementById('userModal').style.display = 'block';
                });
        }

        function deleteUser(id) {
            if (confirm('Are you sure you want to delete this user?')) {
                fetch(`/api/users/${id}`, { method: 'DELETE' })
                    .then(response => response.json())
                    .then(result => {
                        if (result.success) {
                            alert('User deleted successfully!');
                            loadUsers();
                            loadDashboard();
                        } else {
                            alert('Error: ' + result.error);
                        }
                    });
            }
        }

        document.getElementById('userForm').addEventListener('submit', function(e) {
            e.preventDefault();
            
            const formData = new FormData(this);
            const data = Object.fromEntries(formData);
            
            const method = currentEditId ? 'PUT' : 'POST';
            const url = currentEditId ? `/api/users/${currentEditId}` : '/api/users';
            
            fetch(url, {
                method: method,
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(data)
            })
            .then(response => response.json())
            .then(result => {
                if (result.success) {
                    alert(currentEditId ? 'User updated successfully!' : 'User created successfully: ' + result.user_id);
                    closeModal('userModal');
                    loadUsers();
                    loadDashboard();
                } else {
                    alert('Error: ' + result.error);
                }
            });
        });

        // Generate reports
        function generateReport(type) {
            window.open(`/api/reports/${type}`, '_blank');
        }

        // Modal functions
        function closeModal(modalId) {
            document.getElementById(modalId).style.display = 'none';
            currentEditId = null;
            currentEditType = null;
        }

        window.onclick = function(event) {
            if (event.target.classList.contains('modal')) {
                event.target.style.display = 'none';
                currentEditId = null;
                currentEditType = null;
            }
        }

        // Initialize map
        function initMap() {
            if (map) return;
            
            map = L.map('map').setView([24.7136, 46.6753], 6);
            
            L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
                attribution: '© OpenStreetMap contributors'
            }).addTo(map);
            
            map.on('click', function(e) {
                map.eachLayer(function (layer) {
                    if (layer instanceof L.Marker) {
                        map.removeLayer(layer);
                    }
                });
                
                L.marker([e.latlng.lat, e.latlng.lng]).addTo(map);
                
                document.getElementById('latInput').value = e.latlng.lat.toFixed(6);
                document.getElementById('lngInput').value = e.latlng.lng.toFixed(6);
            });
        }
    </script>
</body>
</html>
'''

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)

